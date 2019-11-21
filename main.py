import argparse
import json
import openpyxl
import os
import requests
import sys


parser = argparse.ArgumentParser()
parser.add_argument('token', type=str, help='Токен для Huntflow API')
parser.add_argument('path', type=str, help='Расположение excel файла для разбора')
args = parser.parse_args()

url = 'https://dev-100-api.huntflow.ru'
resume_load_file = 'resume_load.txt'
path_to_file = os.path.join(args.path)
token = args.token

headers = {
    'Authorization': f'Bearer {token}',
    'User-Agent': 'App/1.0 (incaseoffire@example.com)',
}

# функция создает готовый для загрузки список словарей из данных файла excel и путей к резюме
def load_data_from_file(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['Лист1']
    # проверить существование файла
    if os.path.isfile(resume_load_file):
        with open(resume_load_file, 'r', encoding='utf8') as f:
            rows = int(f.readline())
    rows = sheet.max_row
    cols = sheet.max_column
    candidates_list = []
    for row in range(2, rows + 1):
        candidate_fields = {}
        for col in range(1, cols + 1):
            head_cell = sheet.cell(row=1, column=col)
            cell = sheet.cell(row=row, column=col)
            candidate_fields[head_cell.value] = str(cell.value).strip()
        candidates_list.append(candidate_fields)
    return candidates_list


# получение account ID
def fetch_account_id():
    try:
        resp = requests.get(
            url=url + '/accounts',
            headers=headers
        ).json()
        return resp['items'][0]['id']
    except:
        print('Ошибка при получении account_id')
        sys.exit(1)


# добавление кандидата в БД
def add_candidat_to_db(candidat):
    # разбиваем ФИО на 3 отдельных поля: фамилия, имя, отчество
    fio_list = candidat['ФИО'].split()
    if len(fio_list) == 3:
        patronymic = fio_list[2].strip()
    else:
        patronymic = ''

    request_dict = {
        'last_name': fio_list[0].strip(),
        'first_name': fio_list[1].strip(),
        'middle_name': patronymic,
        'position': candidat['Должность'],
        'money': candidat['Ожидания по ЗП'],
    }
    if candidat['ИД_Фото']:
        request_dict['photo'] = candidat['ИД_Фото']

    if candidat['ИД_Файла']:
        request_dict['externals'] = [
            {
                'files': [
                    {
                        'id': candidat['ИД_Файла']
                    },
                ],
            }
        ]
    request_dict = json.dumps(request_dict)
    account_id = fetch_account_id()
    resp = requests.post(
        url=url + f'/account/{account_id}/applicants',
        headers=headers,
        data=request_dict
    ).json()
    candidat['ИД_Резюме'] = resp['id']


# установка статуса кандидата в БД
def set_status_for_candidate(candidat):
    request_dict = {
        'vacancy': candidat['ИД_Вакансии'],
        'status': candidat['ИД_Статуса'],
        'comment': candidat['Комментарий'],
    }
    request_dict = json.dumps(request_dict)
    account_id = fetch_account_id()
    applicant_id = candidat['ИД_Резюме']
    resp = requests.post(
        url=url + f'/account/{account_id}/applicants/{applicant_id}/vacancy',
        headers=headers,
        data=request_dict
    ).json()


# получение списка вакансий
def fetch_vacancies_list():
    account_id = fetch_account_id()
    try:
        resp = requests.get(
            url=url + f'/account/{account_id}/vacancies',
            headers=headers,
        ).json()
        return resp['items']
    except:
        print('Ошибка при получении списка вакансий')
        sys.exit(1)


# получение списка статусов
def fetch_statuses_list():
    account_id = fetch_account_id()
    try:
        resp = requests.get(
            url=url + f'/account/{account_id}/vacancy/statuses',
            headers=headers,
        ).json()
        return resp['items']
    except:
        print('Ошибка при получении списка статусов')
        sys.exit(1)


# добавляем ID вакансии к кандидату
def add_vacancy_id_to_candidat(vacancies, candidat):
    for vacancy in vacancies:
        if candidat['Должность'] == vacancy['position']:
            candidat['ИД_Вакансии'] = vacancy['id']
            break


# добавляем ID статуса к кандидату
def add_status_id_to_candidat(statuses, candidat):
    for status in statuses:
        if candidat['Статус'] == status['name']:
            candidat['ИД_Статуса'] = status['id']
            break


def add_resume_path_to_candidat(candidat):
    folder = os.path.join(os.path.dirname(path_to_file), candidat['Должность'])
    for element in os.scandir(folder):
        if element.is_file():
            if candidat['ФИО'] in element.name:
                candidat['Путь_к_резюме'] = os.path.join(folder, element.name)
            candidat['Путь_к_резюме'] = None


# загрузка резюме
def add_resume_to_db(candidate):
    file_id = None
    photo_id = None
    if candidat['Путь_к_резюме'] is None:
        candidate['ИД_Файла'] = file_id
        candidate['ИД_Фото'] = photo_id
        return
    headers_local = headers.copy()
    headers_local['X-File-Parse'] = 'true'
    file = {'file': open(candidate['Путь_к_резюме'], 'rb')}
    account_id = fetch_account_id()
    try:
        resp = requests.post(
            url=url + f'/account/{account_id}/upload',
            headers=headers_local,
            files=file,
        )
        if resp.status_code == 200:
            resp == resp.json()
            file_id = resp['id']
            photo_id = resp['photo']['id'] if resp['photo'] else None
        candidate['ИД_Файла'] = file_id
        candidate['ИД_Фото'] = photo_id
    except:
        print('Ошибка при загрузке фала резюме')
        sys.exit(1)


# run
print(f'Получение справочников')
vacancies = fetch_vacancies_list()
statuses = fetch_statuses_list()

print('Загрузка данных из файла')
candidats_list = load_data_from_file(path_to_file)

print(f'Загрузка списка кандидатов')
for idx, candidat in enumerate(candidats_list):
    try:
        add_vacancy_id_to_candidat(vacancies, candidat)
        add_status_id_to_candidat(statuses, candidat)
        add_resume_path_to_candidat(candidat)
        add_resume_to_db(candidat)
        add_candidat_to_db(candidat)
        set_status_for_candidate(candidat)
        print(f' - "{candidat["ФИО"]}" успешно загружен')
    except:
        # в случае ошибки записать номер строки в файл
        print(f'Ошибка при загрузке кандидатов')
        with open(resume_load_file, 'w', encoding='utf8') as f:
            f.write(str(idx))

print('Загрузка завершена успешно')
